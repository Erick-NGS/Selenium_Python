from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# lista_items = ['Processador gamer Intel Core i5-10400F BX8070110400F de 6 núcleos e 4.3GHz de frequência', 'Placa Mãe Asus Para Intel 1700 Z690 Plus D4 Tuf 4ddr4 Atx', 'Placa De Video Mancer Rx 5500 Xt Streaky, 8gb, Gddr6 128 Bit', 'Memória RAM NB BLACK color preto 16GB 1 UP Gamer UP3200', 'Disco Sólido Interno Kingston Skc600/512g 512gb Preto Cor Preto', 'Fonte de alimentação para PC Corsair CV Series CV550 550W black 100V/240V', 'Gabinete Gamer Cooler Master Elite 300 Lateral Vidro Preto', 'Teclado Corsair K55 Rgb Multicolor Led', 'Mouse Logitech G G Series G502 Hero preto', 'Monitor Acer 21.5 Zero Frame Radeon Hdmi Ea220q Hbi', 'Headset Gamer Para Consoles E Pc Driver 40mm Quantum 100 Preto Jbl']
# lista_preco = ['720', '1668.90', '949.99', '187.68', '293', '416.74', '328.30', '725', '269', '489', '150']
# lista_links = ['https://www.mercadolivre.com.br/processador-gamer-intel-core-i5-10400f-bx8070110400f-de-6-nucleos-e-43ghz-de-frequncia/p/MLB15915881?pdp_filters=category:MLB1648#searchVariation=MLB15915881&position=2&search_layout=grid&type=product&tracking_id=f4e1fe36-0703-4c18-b222-a180fc41fbd5', 'https://www.mercadolivre.com.br/placa-me-asus-para-intel-1700-z690-plus-d4-tuf-4ddr4-atx/p/MLB22623404?pdp_filters=category:MLB1692#searchVariation=MLB22623404&position=3&search_layout=grid&type=product&tracking_id=fbb7e1d6-d69c-4bab-8c94-308574b4a509', 'https://www.mercadolivre.com.br/placa-de-video-mancer-rx-5500-xt-streaky-8gb-gddr6-128-bit/p/MLB27458201?pdp_filters=category:MLB1658#searchVariation=MLB27458201&position=2&search_layout=grid&type=product&tracking_id=2509c67f-27a0-43a9-82b1-2c40bc8f8328', 'https://www.mercadolivre.com.br/memoria-ram-nb-black-color-preto-16gb-1-up-gamer-up3200/p/MLB22736028?pdp_filters=category:MLB1648#searchVariation=MLB22736028&position=1&search_layout=grid&type=product&tracking_id=2b7e2340-ec32-4a22-ab84-bd49af5de659', 'https://www.mercadolivre.com.br/disco-solido-interno-kingston-skc600512g-512gb-preto-cor-preto/p/MLB24356923?pdp_filters=category:MLB1672#searchVariation=MLB24356923&position=1&search_layout=grid&type=product&tracking_id=a3288598-3937-4dc2-8479-dd0ceb1cfc5e', 'https://www.mercadolivre.com.br/fonte-de-alimentaco-para-pc-corsair-cv-series-cv550-550w-black-100v240v/p/MLB15556690?pdp_filters=category:MLB6777#searchVariation=MLB15556690&position=1&search_layout=grid&type=product&tracking_id=2383f355-f5b0-4808-9b68-77577fc113ff', 'https://produto.mercadolivre.com.br/MLB-4322698874-gabinete-gamer-cooler-master-elite-300-lateral-vidro-preto-_JM#position=2&search_layout=grid&type=item&tracking_id=7c20d6ae-6bb0-4cc9-8f6d-cf9c9f9721d7', 'https://www.mercadolivre.com.br/teclado-corsair-k55-rgb-multicolor-led/p/MLB15172228?pdp_filters=category:MLB1648#searchVariation=MLB15172228&position=2&search_layout=grid&type=product&tracking_id=0b0a2fa9-bfd8-4579-b4c2-ca470e2bb937', 'https://www.mercadolivre.com.br/mouse-logitech-g-g-series-g502-hero-preto/p/MLB12866864?pdp_filters=category:MLB1714#searchVariation=MLB12866864&position=2&search_layout=grid&type=product&tracking_id=6df8d7d4-aad4-464f-be57-c97344d5de7e', 'https://www.mercadolivre.com.br/monitor-acer-215-zero-frame-radeon-hdmi-ea220q-hbi/p/MLB24600090?pdp_filters=category:MLB99245#searchVariation=MLB24600090&position=2&search_layout=grid&type=product&tracking_id=bc651048-aea5-4184-bbd0-bb866f8cf63d', 'https://www.mercadolivre.com.br/headset-gamer-para-consoles-e-pc-driver-40mm-quantum-100-preto-jbl/p/MLB16239943?pdp_filters=item_id:MLB1856964892#is_advertising=true&searchVariation=MLB16239943&position=1&search_layout=stack&type=pad&tracking_id=f5fba67c-39b8-45d3-be38-f7d4ea92fbda&is_advertising=true&ad_domain=VQCATCORE_LST&ad_position=1&ad_click_id=NmU1ZWJhYzgtMTkyYS00NDRkLWJkOWYtN2FiNDA4NjZlZWQ3']
# lista_orcamento = [6198.51]

def criar_tabela(itens, precos, links, orcamento_total):
    ws.title = "Orçamento - PC"
    ws["A1"] = "Itens"
    ws["B1"] = "Preços"
    ws["C1"] = "Links"
    ws["D1"] = "Orçamento"

    index = 2
    for item in itens:
        ws.cell(column=1, row=index, value=item)
        index+=1

    index = 2
    for preco in precos:
        preco = str(preco)
        preco = preco.replace(".", ",")
        ws.cell(column=2, row=index, value=f"R${preco}")
        index+=1

    index = 2
    for link in links:
        ws.cell(column=3, row=index, value=link)
        index+=1

    index = 2
    for orcamento in orcamento_total:
        ws.cell(column=4, row=index, value=orcamento)
        index+=1


    wb.save("Report.xlsx")

# criar_tabela(lista_item, lista_preco, lista_link, lista_orcamento)